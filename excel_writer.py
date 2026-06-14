import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


COLUMNS = ['แพลตฟอร์ม', 'วันที่', 'หมายเลขใบเสร็จ', 'ยอดเงิน (บาท)', 'ชื่อเพจ', 'ID บัญชี', 'ID ธุรกรรม', 'โพสต์/แคมเปญ', 'ชื่อไฟล์', 'สถานะ']

COL_WIDTHS = {
    'แพลตฟอร์ม': 14,
    'วันที่': 22,
    'หมายเลขใบเสร็จ': 28,
    'ยอดเงิน (บาท)': 18,
    'ชื่อเพจ': 30,
    'ID บัญชี': 22,
    'ID ธุรกรรม': 46,
    'โพสต์/แคมเปญ': 60,
    'ชื่อไฟล์': 40,
    'สถานะ': 12,
}

FB_BLUE   = '1877F2'
TT_BLACK  = '010101'
HDR_FG    = 'FFFFFF'
ROW_EVEN  = 'EBF3FF'
ROW_ODD   = 'FFFFFF'
ROW_CHECK = 'FFF3CD'
ROW_ERR   = 'FFE0E0'
SUM_BG    = 'FFE0B2'


def _thin_border():
    s = Side(style='thin', color='D0D7DE')
    return Border(left=s, right=s, top=s, bottom=s)


def _apply_header(ws, platform='facebook'):
    color = FB_BLUE if platform == 'facebook' else TT_BLACK
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(name='Arial', bold=True, color=HDR_FG, size=11)
        cell.fill = PatternFill('solid', start_color=color)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _thin_border()
    ws.row_dimensions[1].height = 28


def _apply_row(ws, row_idx: int, data: dict):
    status = data.get('สถานะ', 'OK')
    if status == 'OCR_NEEDED':
        bg = ROW_ERR
    elif status == 'CHECK':
        bg = ROW_CHECK
    else:
        bg = ROW_EVEN if row_idx % 2 == 0 else ROW_ODD

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        val = data.get(col_name, '')
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = Alignment(vertical='center', wrap_text=(col_name == 'โพสต์/แคมเปญ'))
        cell.border = _thin_border()
        if col_name == 'ยอดเงิน (บาท)' and isinstance(val, (int, float)):
            cell.number_format = '#,##0.00'


def _apply_sum_row(ws, data_rows: int, platform: str):
    sum_row = data_rows + 2
    amount_col = COLUMNS.index('ยอดเงิน (บาท)') + 1
    label_col = amount_col - 1

    ws.cell(row=sum_row, column=label_col, value='รวมทั้งหมด').font = Font(name='Arial', bold=True, size=11)
    sum_cell = ws.cell(row=sum_row, column=amount_col)
    sum_cell.value = f'=SUM({get_column_letter(amount_col)}2:{get_column_letter(amount_col)}{data_rows + 1})'
    sum_cell.number_format = '#,##0.00'
    sum_cell.font = Font(name='Arial', bold=True, size=11)

    color = FB_BLUE if platform == 'facebook' else TT_BLACK
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=sum_row, column=col_idx)
        cell.fill = PatternFill('solid', start_color=SUM_BG)
        cell.border = _thin_border()


def create_excel(rows_fb: list, rows_tt: list, rows_err: list, output_path: str):
    wb = openpyxl.Workbook()

    def fill_sheet(ws, rows, platform):
        _apply_header(ws, platform)
        for i, row in enumerate(rows, start=2):
            _apply_row(ws, i, row)
        if rows:
            _apply_sum_row(ws, len(rows), platform)
        # column widths
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(col_name, 18)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    # Sheet 1: Facebook
    ws_fb = wb.active
    ws_fb.title = 'Facebook'
    fill_sheet(ws_fb, rows_fb, 'facebook')

    # Sheet 2: TikTok
    ws_tt = wb.create_sheet('TikTok')
    fill_sheet(ws_tt, rows_tt, 'tiktok')

    # Sheet 3: Summary
    ws_sum = wb.create_sheet('สรุปรวม')
    ws_sum['A1'] = 'สรุปค่าโฆษณา'
    ws_sum['A1'].font = Font(name='Arial', bold=True, size=14)
    ws_sum['A3'] = 'แพลตฟอร์ม'
    ws_sum['B3'] = 'จำนวนใบ'
    ws_sum['C3'] = 'ยอดรวม (บาท)'
    for cell in [ws_sum['A3'], ws_sum['B3'], ws_sum['C3']]:
        cell.font = Font(name='Arial', bold=True, color=HDR_FG)
        cell.fill = PatternFill('solid', start_color='333333')
        cell.alignment = Alignment(horizontal='center')
    ws_sum['A4'] = 'Facebook'
    ws_sum['B4'] = len([r for r in rows_fb if r.get('สถานะ') == 'OK'])
    ws_sum['C4'] = f'=SUM(Facebook!{get_column_letter(COLUMNS.index("ยอดเงิน (บาท)")+1)}2:{get_column_letter(COLUMNS.index("ยอดเงิน (บาท)")+1)}{len(rows_fb)+1})'
    ws_sum['A5'] = 'TikTok'
    ws_sum['B5'] = len([r for r in rows_tt if r.get('สถานะ') == 'OK'])
    ws_sum['C5'] = f'=SUM(TikTok!{get_column_letter(COLUMNS.index("ยอดเงิน (บาท)")+1)}2:{get_column_letter(COLUMNS.index("ยอดเงิน (บาท)")+1)}{len(rows_tt)+1})'
    ws_sum['A6'] = 'รวมทั้งหมด'
    ws_sum['B6'] = '=B4+B5'
    ws_sum['C6'] = '=C4+C5'
    for row in [ws_sum['A6'], ws_sum['B6'], ws_sum['C6']]:
        row.font = Font(name='Arial', bold=True)
        row.fill = PatternFill('solid', start_color=SUM_BG)
    for col in ['A', 'B', 'C']:
        ws_sum.column_dimensions[col].width = 22
    ws_sum['C4'].number_format = '#,##0.00'
    ws_sum['C5'].number_format = '#,##0.00'
    ws_sum['C6'].number_format = '#,##0.00'
    ws_sum['A8'] = f'สร้างเมื่อ: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws_sum['A8'].font = Font(name='Arial', color='888888', size=9)

    # Sheet 4: Error log
    if rows_err:
        ws_err = wb.create_sheet('ต้องตรวจสอบ')
        fill_sheet(ws_err, rows_err, 'facebook')

    wb.save(output_path)
    return output_path

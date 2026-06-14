"""
excel_merger.py — รวม Excel หลายไฟล์ (หลายเพจ) เป็นไฟล์เดียว
พร้อม sheet รายบิล + sheet ยอดรวมรายวัน
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from collections import defaultdict
from pathlib import Path

# ── styles ────────────────────────────────────────────────────
HDR_MERGED  = '1A2340'   # header รายบิล
HDR_DAILY   = '0F3D2E'   # header รายวัน
ROW_EVEN    = 'F0F4FF'
ROW_ODD     = 'FFFFFF'
SUM_BG      = 'FFE0B2'
DAILY_TOTAL = 'D1FAE5'
FB_BLUE     = '1877F2'
TT_PINK     = 'FE2C55'

BILL_COLS = [
    ('แพลตฟอร์ม',       12),
    ('วันที่',           14),
    ('หมายเลขใบเสร็จ',   28),
    ('ยอดเงิน (บาท)',    16),
    ('ชื่อเพจ',          28),
    ('ID บัญชี',         22),
    ('โพสต์/แคมเปญ',    50),
    ('สถานะ',            10),
]
DAILY_COLS = [
    ('วันที่',           14),
    ('จำนวนใบ',         12),
    ('ยอดรวม (บาท)',    18),
]

def _tb():
    s = Side(style='thin', color='D0D7DE')
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_cell(cell, text, bg, fg='FFFFFF', sz=10):
    cell.value = text
    cell.font = Font(name='Arial', bold=True, color=fg, size=sz)
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = _tb()

def _data_cell(cell, val, even, fmt=None, bold=False, color=None):
    cell.value = val
    cell.font = Font(name='Arial', size=10, bold=bold, color=color or '000000')
    cell.fill = PatternFill('solid', start_color=ROW_EVEN if even else ROW_ODD)
    cell.alignment = Alignment(vertical='center')
    cell.border = _tb()
    if fmt:
        cell.number_format = fmt

def _plat_badge(ws, row, col, platform):
    cell = ws.cell(row=row, column=col, value=platform)
    color = FB_BLUE if platform == 'Facebook' else TT_PINK
    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color=color)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = _tb()


def _parse_date_sort_key(date_str: str):
    """แปลง DD/MM/YYYY → tuple สำหรับ sort"""
    try:
        parts = date_str.strip().split('/')
        return (int(parts[2]), int(parts[1]), int(parts[0]))
    except Exception:
        return (9999, 99, 99)


def read_excel_rows(xlsx_path: str) -> list[dict]:
    """อ่าน rows จาก Excel ที่สร้างโดยระบบนี้ (sheet Facebook / TikTok)"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    rows = []
    for sheet_name in ['Facebook', 'TikTok']:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        for r in range(2, ws.max_row + 1):
            row = {}
            for c, h in enumerate(headers, 1):
                if h:
                    row[h] = ws.cell(r, c).value
            # ข้ามแถวว่างและแถว "รวมทั้งหมด"
            if not row.get('หมายเลขใบเสร็จ') and not row.get('วันที่'):
                continue
            if row.get('หมายเลขใบเสร็จ') == 'รวมทั้งหมด' or row.get('แพลตฟอร์ม') == 'แพลตฟอร์ม':
                continue
            rows.append(row)
    return rows


def merge_excels(source_paths: list[str], extra_rows: list[dict] = None) -> openpyxl.Workbook:
    """
    รวม Excel หลายไฟล์ → Workbook ที่มี:
      Sheet 1: รายบิลทั้งหมด (เรียงตามวันที่ → เพจ)
      Sheet 2: ยอดรวมรายวัน
    """
    # รวม rows จากทุกแหล่ง
    all_rows = []
    for path in source_paths:
        try:
            all_rows.extend(read_excel_rows(path))
        except Exception as e:
            print(f"[WARN] อ่านไม่ได้: {path} — {e}")

    if extra_rows:
        all_rows.extend(extra_rows)

    # deduplicate ด้วย หมายเลขใบเสร็จ
    seen = set()
    unique_rows = []
    for r in all_rows:
        key = str(r.get('หมายเลขใบเสร็จ', '')) + str(r.get('ID ธุรกรรม', ''))
        if key and key not in seen:
            seen.add(key)
            unique_rows.append(r)
        elif not key:
            unique_rows.append(r)

    # เรียงตามวันที่ แล้วตามเพจ
    unique_rows.sort(key=lambda r: (
        _parse_date_sort_key(str(r.get('วันที่', ''))),
        str(r.get('ชื่อเพจ', '')),
    ))

    wb = openpyxl.Workbook()

    # ══ SHEET 1: รายบิลทั้งหมด ══════════════════════════════════
    ws_bill = wb.active
    ws_bill.title = 'รายบิลทั้งหมด'

    # header
    for ci, (col_name, _) in enumerate(BILL_COLS, 1):
        _hdr_cell(ws_bill.cell(1, ci), col_name, HDR_MERGED, sz=11)
    ws_bill.row_dimensions[1].height = 28

    # data rows
    total_amount = 0.0
    for ri, row in enumerate(unique_rows, 2):
        even = ri % 2 == 0
        _plat_badge(ws_bill, ri, 1, row.get('แพลตฟอร์ม', ''))
        _data_cell(ws_bill.cell(ri, 2), row.get('วันที่', ''), even)
        _data_cell(ws_bill.cell(ri, 3), row.get('หมายเลขใบเสร็จ', ''), even,
                   color='1877F2' if row.get('แพลตฟอร์ม') == 'Facebook' else 'FE2C55')
        amt = row.get('ยอดเงิน (บาท)', 0) or 0
        _data_cell(ws_bill.cell(ri, 4), float(amt), even, fmt='#,##0.00', bold=True)
        total_amount += float(amt)
        _data_cell(ws_bill.cell(ri, 5), row.get('ชื่อเพจ', ''), even)
        _data_cell(ws_bill.cell(ri, 6), row.get('ID บัญชี', ''), even)
        _data_cell(ws_bill.cell(ri, 7), row.get('โพสต์/แคมเปญ', ''), even)
        status = row.get('สถานะ', 'OK')
        sc = ws_bill.cell(ri, 8)
        sc.value = status
        sc.font = Font(name='Arial', size=10, bold=True,
                       color='22C55E' if status == 'OK' else 'F59E0B')
        sc.fill = PatternFill('solid', start_color=ROW_EVEN if even else ROW_ODD)
        sc.alignment = Alignment(horizontal='center', vertical='center')
        sc.border = _tb()

    # total row
    total_row = len(unique_rows) + 2
    ws_bill.cell(total_row, 3, 'รวมทั้งหมด').font = Font(name='Arial', bold=True, size=11)
    tc = ws_bill.cell(total_row, 4)
    tc.value = total_amount
    tc.number_format = '#,##0.00'
    tc.font = Font(name='Arial', bold=True, size=12, color='1A2340')
    for ci in range(1, 9):
        c = ws_bill.cell(total_row, ci)
        c.fill = PatternFill('solid', start_color=SUM_BG)
        c.border = _tb()

    # column widths + freeze + filter
    for ci, (_, w) in enumerate(BILL_COLS, 1):
        ws_bill.column_dimensions[get_column_letter(ci)].width = w
    ws_bill.freeze_panes = 'A2'
    ws_bill.auto_filter.ref = f'A1:{get_column_letter(len(BILL_COLS))}{len(unique_rows)+1}'

    # ══ SHEET 2: ยอดรวมรายวัน ════════════════════════════════════
    ws_day = wb.create_sheet('ยอดรวมรายวัน')

    # รวม rows ตามวัน
    daily: dict[str, dict] = {}
    for row in unique_rows:
        d = str(row.get('วันที่', '') or '').strip()
        if not d:
            continue
        if d not in daily:
            daily[d] = {'วันที่': d, 'จำนวนใบ': 0, 'ยอดรวม (บาท)': 0.0, '_sort': _parse_date_sort_key(d)}
        daily[d]['จำนวนใบ'] += 1
        daily[d]['ยอดรวม (บาท)'] += float(row.get('ยอดเงิน (บาท)', 0) or 0)

    daily_rows = sorted(daily.values(), key=lambda x: x['_sort'])

    # header
    for ci, (col_name, _) in enumerate(DAILY_COLS, 1):
        _hdr_cell(ws_day.cell(1, ci), col_name, HDR_DAILY, sz=11)
    ws_day.row_dimensions[1].height = 28

    grand_total = 0.0
    for ri, dr in enumerate(daily_rows, 2):
        even = ri % 2 == 0
        _data_cell(ws_day.cell(ri, 1), dr['วันที่'], even, bold=True)
        _data_cell(ws_day.cell(ri, 2), dr['จำนวนใบ'], even, fmt='#,##0')
        _data_cell(ws_day.cell(ri, 3), dr['ยอดรวม (บาท)'], even, fmt='#,##0.00', bold=True,
                   color='059669')
        grand_total += dr['ยอดรวม (บาท)']

    # grand total
    gt_row = len(daily_rows) + 2
    gt_label = ws_day.cell(gt_row, 1, 'รวมทั้งหมด')
    gt_label.font = Font(name='Arial', bold=True, size=11)
    gt_cnt = ws_day.cell(gt_row, 2, len(unique_rows))
    gt_cnt.font = Font(name='Arial', bold=True, size=11)
    gt_amt = ws_day.cell(gt_row, 3, grand_total)
    gt_amt.number_format = '#,##0.00'
    gt_amt.font = Font(name='Arial', bold=True, size=12, color='059669')
    for ci in range(1, 4):
        c = ws_day.cell(gt_row, ci)
        c.fill = PatternFill('solid', start_color=DAILY_TOTAL)
        c.border = _tb()

    for ci, (_, w) in enumerate(DAILY_COLS, 1):
        ws_day.column_dimensions[get_column_letter(ci)].width = w
    ws_day.freeze_panes = 'A2'
    ws_day.auto_filter.ref = f'A1:C{len(daily_rows)+1}'

    # ── metadata
    ws_day['E1'] = f'สร้างเมื่อ {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws_day['E1'].font = Font(name='Arial', color='888888', size=9)
    ws_day['E2'] = f'รวม {len(unique_rows)} ใบ จาก {len(source_paths)} ไฟล์'
    ws_day['E2'].font = Font(name='Arial', color='888888', size=9)

    return wb
